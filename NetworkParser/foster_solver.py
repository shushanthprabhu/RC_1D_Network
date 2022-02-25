# !/usr/bin/env python3
# coding: utf-8
__author__ = "Shushanth Prabhu"
__email__ = "shushanth@gmail.com"
__version__ = "1.0"

from scipy.optimize import least_squares
from numpy import exp, inf, asarray
from math import sqrt
from openpyxl import Workbook, load_workbook
from NetworkParser import FC_Error_list
from os.path import exists
from os import remove
from os import path

# from matplotlib import pyplot as plt
# from matplotlib.ticker import (MultipleLocator, FormatStrFormatter,
#                                AutoMinorLocator)
from math import log10

# TODO : if exact solution is known can we evaluate max error
# TODO : In Diagfile create Plots
# TODO : Export Circuite to *.asc file

maximum_network_elements = 10
diagnostic_filename = 'Foster_Diagnostic.xlsx'


def check_null_value(trace, loc):
    """
    Algorithm raises an error if First Zth point is not Zero
    :param trace:
    :return: True if start point is not Zero
    """
    if trace[loc] != 0:
        raise FC_Error_list.FosterCauer_Error("Start point is Not Zero")
    else:
        return True


def foster_default_value():
    """
    Property containing all default foster network values
    """
    # LOWER BOUND
    default_foster_lower_bound = 0.00000001
    # UPPER BOUND
    default_foster_upper_bound = 1000000
    # Tolerance of Cost Function ftol
    default_foster_ftol = 0.00000001
    # Tolerance of Independent Functions xtol
    default_foster_xtol = 0.00000001
    # CREATE DIAGONOSTIC
    default_foster_diagnostic = False
    default_cauer_diagnostic = False
    return (
        default_foster_lower_bound, default_foster_upper_bound, default_foster_ftol, default_foster_xtol,
        default_foster_diagnostic, default_cauer_diagnostic)


def read_csv_file(file):
    """
    Function to read a CSV file
    :param file: a CSV File
    :return: List containing data
    """

    if not path.exists(file):
        raise FC_Error_list.FosterCauer_Error("File not found")

    else:
        try:
            input_file = open(file, 'r')
            lines = input_file.readlines()
            if not len(lines) or len(lines) == 1:
                raise FC_Error_list.FosterCauer_Error("No Data in Input File")
            else:
                output_list = []
                for line in lines:
                    split_line = line.split(",")
                    output_line = []
                    for line_item in split_line:
                        line_item.rstrip("\n")
                        try:
                            if float(line_item):
                                output_line.append(float(line_item))
                            if float(line_item) == 0.0:
                                output_line.append(0.0)
                        except Exception:
                            pass
                    if len(output_line) > 0:
                        output_list.append(output_line)

            if check_null_value(output_list[0], 1):
                return output_list
        except IndexError:
            raise FC_Error_list.FosterCauer_Error("Input file not organized correctly, Refer documentation")


def sanity_check(time_trace, input_zth):
    """
    Function to create to do a sanity check of Input values.
    :param time_trace: time plot in s
    :param input_zth: input Zth
    :return: Boolean (True if all conditions are cleared False if not)
    """
    output = True
    # CHECK IF Time Trace and Zth Trace are equal
    if not len(time_trace) == len(input_zth):
        output = False

    # CHECK IF THERE ARE NEGATIVE VALUES IN Zth
    for i in input_zth:
        if i < 0:
            raise FC_Error_list.FosterCauer_Error("Negative Zth in Input file")

    # CHECK IF NEGATIVE TIME SLOP EXISTS
    i = 0
    for i in range(0, len(time_trace) - 1):
        if time_trace[i + 1] < time_trace[i]:
            raise FC_Error_list.FosterCauer_Error("Time trace is Negative")

    # CHECK IF Time TRACE AND Zth TRACE ARE GREATER THAN 3
    if len(time_trace) > 3 & len(input_zth) > 3:
        raise FC_Error_list.FosterCauer_Error("Input data too short")

    return output


def foster_func(tpl, x):
    """
    Function returning the Foster Function value
    :param tpl: Contain the RC Pairs with R Value First and C Value Seconds
    :param x: Time
    :return: Foster Function at Time t in s
    """
    counter = 0
    return_value = 0

    while counter < len(tpl):
        return_value += tpl[counter] * (1 - exp(-x / tpl[counter] / tpl[counter + 1]))
        counter += 2
    return return_value


def error_func(tpl, x, y):
    """
    Function returning the Error between Present value and foster function calculated at x in s
    :param tpl: Contain the RC Pairs with R Value First and C Value Seconds
    :param x: Time point at which value is calculated.
    :param y: Present Guessed value
    :return: Error Function
    """
    return (foster_func(tpl, x) - y)
    # *(foster_func(tpl, x) - y)


def sort_tpl(rc_list):
    """
    Function to Split RC Tuple to Individual lists.
    :param rc_list: Tuple containing RC List together in R and C format
    :return: Two tuple containing R Parameters and C Parameters
    """
    r_list = []
    c_list = []
    counter = 0
    while counter != len(rc_list):
        r_list.append(rc_list[counter])
        c_list.append(rc_list[counter + 1])
        counter += 2
    return r_list, c_list


def solve_foster(time, input_zth, count_elements, default_tuple=None):
    """
    Curve Fitting RC Values
    :param time: time in seconds
    :param input_zth: Zth to match values
    :param count_elements: Maximum number of elements to use.
    :return: tuple containing R and C List.
    """
    initial_guess = (0.1, 0.1) * count_elements
    try:
        foster_lower_bound = default_tuple[0]
        foster_upper_bound = default_tuple[1]
        foster_ftol = default_tuple[2]
        foster_xtol = default_tuple[3]
    except:
        foster_lower_bound = 0.00000001
        foster_upper_bound = inf
        foster_ftol = 1e-08
        foster_xtol = 1e-08

    output = least_squares(error_func, x0=initial_guess, jac='2-point', bounds=(foster_lower_bound, foster_upper_bound),
                           method='trf',
                           ftol=foster_ftol,
                           xtol=foster_xtol, gtol=1e-08, x_scale=1.0, loss='linear', f_scale=1.0, diff_step=None,
                           tr_solver=None,
                           tr_options={}, jac_sparsity=None, max_nfev=None, verbose=0, args=(time, input_zth))

    # output = least_squares(error_func, x0=initial_guess, jac='2-point', bounds=(0, np.inf), method='trf', ftol=1e-08,
    #                        xtol=1e-08, gtol=1e-08, x_scale=1.0, loss='linear', f_scale=1.0, diff_step=None,
    #                        tr_solver=None,
    #                        tr_options={}, jac_sparsity=None, max_nfev=None, verbose=0, args=(time, input_zth))

    tpl_final = output.x
    # sort_tpl(tpl_final)
    # print(tpl_final)
    return tpl_final


def delete_diagnostic_file():
    """
    Delete the diagnostic File writter earlier.
    :return: None
    """
    if exists(diagnostic_filename):
        remove(diagnostic_filename)


def create_diagnostic_header(time_trace, input_z):
    """
    Create a Diagnostic header
    :param time_trace: time in seconds
    :param input_z: input Impedance
    :return: False if Diagnostic file exists True if new file is created.
    """
    if exists(diagnostic_filename):
        return False
    else:
        wb = Workbook()
        sheet = wb['Sheet']
        sheet.title = "Network_Combinations"
        sheet.cell(row=1, column=1, value="Number of Elements")
        sheet.cell(row=2, column=1, value="RC Network")
        sheet.cell(row=3, column=1, value="Error")
        sheet.cell(row=4, column=1, value="Error Trace")

        sheet_zth = wb.create_sheet(title="Zth")
        sheet_zth.cell(row=1, column=1, value="Time")
        sheet_zth.cell(row=1, column=2, value="Input Zth")
        r = 2
        for i in time_trace:
            sheet_zth.cell(row=r, column=1, value=i)
            r += 1
        r = 2
        for i in input_z:
            sheet_zth.cell(row=r, column=2, value=i)
            r += 1

        wb.save(diagnostic_filename)
        wb.close()
        return True


def write_diagnostic(item, time_trace, input_z):
    """
    Write Diagnostic file
    :param item:
    :param time_trace:
    :param input_z:
    :return:
    """
    if create_diagnostic_header(time_trace, input_z):
        workbook = load_workbook(diagnostic_filename)
        worksheet = workbook["Network_Combinations"]
        col = item[0] + 1
        worksheet.cell(row=1, column=col, value=item[0])
        worksheet.cell(row=2, column=col, value=str(item[1]))
        worksheet.cell(row=3, column=col, value=item[2])
        r = 4
        for i in item[3]:
            worksheet.cell(row=r, column=col, value=i)
            r += 1

        r = 2
        zth_worksheet = workbook["Zth"]
        col = item[0] + 2
        zth_worksheet.cell(row=1, column=col, value=str(item[0] * 2) + " Elements")
        z_curve = foster_func(item[1], time_trace)
        for i in z_curve:
            zth_worksheet.cell(row=r, column=col, value=i)
            r += 1
        workbook.save(diagnostic_filename)
        workbook.close()


def create_foster_network(time, input_zth, default_tuple, max_elements=None):
    """
    Function to create a Foster Network
    :param time: time in seconds
    :param input_zth: zth
    :param diagnostic: Toggle to turn on Diagnostic
    :param max_elements: Max number of allowed elements
    :return: RC network output combined
    """
    diagnostic = default_tuple[4]
    maximum_elements = max_elements or maximum_network_elements
    delete_diagnostic_file()
    element_count = 1
    network_combination = []
    while element_count != maximum_elements:
        rc_network_list = solve_foster(time, input_zth, element_count, default_tuple)
        error = error_func(rc_network_list, time, input_zth)
        sum_error_square = sqrt(sum(map(lambda i: i * i, error)))
        element = (element_count, rc_network_list, sum_error_square, error)
        network_combination.append(element)

        if diagnostic:
            write_diagnostic(element, time, input_zth)
        element_count += 1

    minimal = 100000000000
    minimal_location = -1
    for item in network_combination:
        if minimal > item[2]:
            minimal = item[2]
            minimal_location = network_combination.index(item)

    output_network = network_combination[minimal_location][1]
    return output_network


def get_foster_network(file, default_values):
    """
    Function to create a Foster network from a CSV file
    :param file: CSV File
    :return: RC Network
    """
    if file.strip() or not file.isspace():
        zth_trace = read_csv_file(file)
        zth_trace = asarray(zth_trace)
        default_values = default_values

        if sanity_check(zth_trace[:, 0], zth_trace[:, 1]):
            try:
                network_rc = create_foster_network(zth_trace[:, 0], zth_trace[:, 1], default_values)
                return network_rc
            except IndexError:
                raise FC_Error_list.FosterCauer_Error("Input file not organized correctly, Refer documentation")
            except Exception:
                raise FC_Error_list.FosterCauer_Error("Random Error \n Please contact the Author")


def sort_rc_list(network):
    """
    Sort RC network to R List and C List
    :param network:
    :return: tuple with R List and C List
    """
    list_1 = []
    list_2 = []
    item = 0
    while item < len(network):
        list_1.append(network[item])
        list_2.append(network[item + 1])
        item += 2
    return (list_1, list_2)


def percentage_error(abs_error, base_value):
    """
    Function to Calculate the percentage error
    :param abs_error: Absolute Error
    :param base_value: Base Value
    :return: percentage error
    """
    # print("Absolute Error-")
    # print(abs_error)
    # print("Base Value-")
    # print(base_value)
    error = []
    error.append(0)
    counter = 1
    while counter != len(abs_error):
        error.append((abs_error[counter]) / base_value[counter])
        counter += 1
    return error


def list_log10(list):
    """
    Returns list with log10
    :param list:
    :return:
    """
    final_list = []
    for item in list:
        final_list.append(log10(item))
    return final_list


# def plot_zth(file, network):
#     """
#     Function to plot the Zth Curve with Absolute Error
#     :param file:
#     :param network:
#     :return:
#     """
#     zth_input = read_csv_file(file)
#     zth_input = asarray(zth_input)
#
#     plt.figure()
#
#     plt.subplot(311)
#     log_time = list_log10(zth_input[:, 0])
#     plt.semilogx(log_time, zth_input[:, 1])
#     plt.semilogx(log_time, foster_func(network, zth_input[:, 0]))
#     plt.xlabel("Time [s]")
#     plt.ylabel("Voltage [V]")
#     plt.grid(True)
#
#     plt.subplot(323)
#     error = error_func(network, zth_input[:, 0], zth_input[:, 1])
#     percent_error = percentage_error(error, zth_input[:, 1])
#     plt.plot(zth_input[:, 0], percent_error)
#     # plt.xlabel("Time [s]")
#     # plt.ylabel("Percentage Error")
#     plt.grid(True)
#
#     plt.subplot(324)
#     error = error_func(network, zth_input[:, 0], zth_input[:, 1])
#     percent_error = percentage_error(error, zth_input[:, 1])
#     plt.plot(zth_input[:, 0], percent_error)
#     plt.xlabel("Time [s]")
#     plt.ylabel("Percentage Error")
#     plt.grid(True)
#
#     plt.show()


def sort_text_file(filename):
    if not path.exists(filename):
        raise FC_Error_list.FosterCauer_Error("File not found")

    else:
        try:
            input_file = open(filename, 'r')
            line = input_file.readline()
            cnt = 1

            while line:
                # print("Line {}: {}".format(cnt, line.strip()))
                line = input_file.readline()
                cnt += 1
            factor = 1 + (cnt // 1048576)
            input_file = open(filename, 'r')
            line = input_file.readline()
            cnt = 1
            final_list = []

            while line:
                if (cnt % factor == 0):
                    final_list.append(line)
                line = input_file.readline()
                cnt += 1

            with open('reduced_file.txt', 'w+') as f:
                for item in final_list:
                    f.write("%s" % item)



        except IndexError:
            raise FC_Error_list.FosterCauer_Error("Input file not organized correctly, Refer documentation")


if __name__ == "__main__":
    csv_file = r"C:\Users\ocu01756\OneDrive - Osram-Continental GmbH\Documents\personal\Papers\RC_Network\FosterNetwork\ExampleCircuit\SW_Validation_2\0.5P"
    csv_file += r"\Book1.csv"
    # csv_file = r"C:\Users\ocu01756\OneDrive - Osram-Continental GmbH\Documents\python_projects\Network_Solver\Test"
    # csv_file += r"\Network_Modelling_Input.csv"

    rc_network = get_foster_network(csv_file, foster_default_value())
    # rc_network = [6.42295735e-01, 1.25570161e-01, 3.99380526e-01, 1.39386358e-04, 6.62249220e+00, 8.46544983e-04,
    #               1.89898094e+01, 5.21140458e+01,
    #               3.55159329e-03, 2.75982519e+02]
    print(rc_network)

    r_list, c_list = sort_rc_list(rc_network)
    print(r_list)
    print(c_list)
    # plot_zth(csv_file, rc_network)

    # text_file = r"C:\Users\ocu01756\OneDrive - Osram-Continental GmbH\Documents\personal\Papers\RC_Network\FosterNetwork\ExampleCircuit\SW_Validation_2\0.5P"
    # text_file += r"\Foster_0.5_PWM1.txt"
    # sort_text_file(text_file)
